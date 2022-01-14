import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from datetime import datetime

def main():

    usuario = str(input("Qual o seu usuário?\n"))
    n_Obra = str(input("\nQual o número da obra?\n"))
    qtd_encarregado = int(input("\nQual o nº de encarregados nesta Obra?\n"))

    print("\nEncarregados Regional Leste:\n\n1 - MARCONDES")
    print("2 - LEUDSON")
    print("3 - CRISTIANO SA")
    print("4 - VALTEIR")
    print("5 - EDENILTON")
    print("6 - FRANCISCO")
    print("7 - CRISTIANO BATISTA")
    print("8 - LUGRECIO\n")
    #print("Encarregados Regional Oeste:\n1 - ")

    Enc = ['MARCONDES', 'LEUDSON', 'CRISTIANO SA', 'VALTEIR', 'EDENILTON', 'FRANCISCO', 'CRISTIANO BATISTA', 'LUGRECIO']
    Encarregados = []
    contador_e = 0

    while contador_e != qtd_encarregado:

        Encarregado = int(input("Qual o encarregado?\n"))
        Encarregados.append(Encarregado)
        contador_e = contador_e + 1

    
    print("\Fiscais das Obras:\n\n1 - CRISPIM")
    print("2 - EDIMILSON")
    print("3 - GIDEONE")
    print("4 - MARIO")
    print("5 - CAJÉ")
    print("6 - WILLMENNES")

    Fiscal = int(input("\nQual o fiscal desta Obra?\n"))
    Fiscais = ['CRISPIM', 'EDIMILSON', 'GIDEONE', 'MARIO', 'CAJÉ', 'WILLMENNES']
    
    Planilha = load_workbook("C:\\Users\\{}\\Documents\\Documentos Obras\\PLANILHA-DE-MÃO-DE-OBRA-CONTROL.xlsx".format(usuario, n_Obra))
    SIAGO = load_workbook("C:\\Users\\{}\\Downloads\\{}.xlsx".format(usuario, n_Obra))

    LM = Planilha["TABELA PADRAO LM - LESTE-OESTE"]
    LV = Planilha["TABELA PADRAO LV - LESTE-OESTE"]
    Aba_SIAGO = SIAGO.active
 
    img = openpyxl.drawing.image.Image('CONTROL.jfif') 
    img.anchor = 'H2'
    img.height = 78
    img.width = 203
    LM.add_image(img)

    img2 = openpyxl.drawing.image.Image('CONTROL.jfif') 
    img2.anchor = 'H2'
    img2.height = 78
    img2.width = 173
    LV.add_image(img2)

    for celulaSIAGO in Aba_SIAGO['L']:  
        linhaSIAGO = celulaSIAGO.row
        verificação = Aba_SIAGO["L{}".format(linhaSIAGO)].value

        if "LINHA VIVA" in verificação:
            codigoSIAGO = Aba_SIAGO["F{}".format(linhaSIAGO)].value
            serviço = Aba_SIAGO["G{}".format(linhaSIAGO)].value
            qtdserviço = Aba_SIAGO["H{}".format(linhaSIAGO)].value
            
            for celulaLV in LV['B']:
                linhaLV = celulaLV.row
                codigoLV = LV["B{}".format(linhaLV)].value

                if codigoLV == codigoSIAGO:
                    if serviço == "R":
                        LV["G{}".format(linhaLV)] = qtdserviço
                    
                    elif codigoLV == 1052:
                        LV["E{}".format(linhaLV)] = qtdserviço
                    
                    elif codigoLV == 1053:
                        LV["E{}".format(linhaLV)] = qtdserviço
                    
                    elif codigoLV == 1056:
                        LV["E{}".format(linhaLV)] = qtdserviço

                    elif codigoLV == 1057:
                        LV["E{}".format(linhaLV)] = qtdserviço

                    elif codigoLV == 1058:
                        LV["E{}".format(linhaLV)] = qtdserviço

                    elif codigoLV == 1059:
                        LV["E{}".format(linhaLV)] = qtdserviço

                    else:
                        LV["D{}".format(linhaLV)] = qtdserviço

                else:
                    print("Aguarde... Escrevendo...")
        
        else:
            codigoSIAGO = Aba_SIAGO["F{}".format(linhaSIAGO)].value

            if codigoSIAGO == 1119:
                codigoSIAGO = 4633
            elif codigoSIAGO == 1120:
                codigoSIAGO = 4632
            else:
                codigoSIAGO = codigoSIAGO
                
            serviço = Aba_SIAGO["G{}".format(linhaSIAGO)].value
            qtdserviço = Aba_SIAGO["H{}".format(linhaSIAGO)].value
            
            for celulaLM in LM['B']:
                linhaLM = celulaLM.row
                codigoLM = LM["B{}".format(linhaLM)].value

                if codigoLM == codigoSIAGO:
                    if serviço == "R":
                        LM["G{}".format(linhaLM)] = qtdserviço

                    elif codigoLM == 1050:
                        LM["E{}".format(linhaLM)] = qtdserviço
                    
                    elif codigoLM == 1051:
                        LM["E{}".format(linhaLM)] = qtdserviço
                    
                    elif codigoLM == 1054:
                        LM["E{}".format(linhaLM)] = qtdserviço

                    elif codigoLM == 1055:
                        LM["E{}".format(linhaLM)] = qtdserviço

                    elif codigoLM == 1058:
                        LM["E{}".format(linhaLM)] = qtdserviço

                    elif codigoLM == 1059:
                        LM["E{}".format(linhaLM)] = qtdserviço
                    
                    else:
                        LM["D{}".format(linhaLM)] = qtdserviço

                else:
                    print("Aguarde... Escrevendo...")

    LM["E4"] = datetime.today().strftime("%d/%m/%Y")
    LV["E4"] = datetime.today().strftime("%d/%m/%Y")

    LM["C6"] = n_Obra
    LV["C6"] = n_Obra

    LM["G6"] = Fiscais[Fiscal-1]
    LV["G6"] = Fiscais[Fiscal-1]


    if len(Encarregados)>1:

        responsável = ""
        equipe = ""

        for i in Encarregados:
            if int(i) < 7:
                
                responsável += "{}, ".format(Enc[i-1])
                equipe += "{}, ".format(Enc[i])
                LM["C4"] = responsável
                LM["C5"] = equipe
            
            elif int(i) >= 7:
                responsável += "{}, ".format(Enc[i-1])
                equipe += "{}, ".format(Enc[i])
                LV["C4"] = responsável
                LV["C5"] = equipe
    
    else:
        if int(Encarregados[0]) < 7:
            LM["C4"] = Enc[Encarregados[0]-1]
        else:
            LV["C4"] = Enc[Encarregados[0]-1]

    print("Planilha finalizada e salva.")

    Planilha.save("C:\\Users\\{}\\OneDrive\\Documentos\\Medição\\{}\\Mão de Obra {}.xlsx".format(usuario,n_Obra,n_Obra))

main()