from openpyxl import Workbook, load_workbook

def main():

    usuario = str(input("Qual o seu usuário?\n"))
    n_Obra = str(input("\nQual o número da obra?\n"))

    Planilha = load_workbook("C:\\Users\\{}\\Documents\\Documentos Obras\\PLANILHA-DE-MÃO-DE-OBRA-CONTROL.xlsx".format(usuario, n_Obra))
    SIAGO = load_workbook("C:\\Users\\{}\\Downloads\\{}.xlsx".format(usuario, n_Obra))

    LM = Planilha["TABELA PADRAO LM - LESTE-OESTE"]
    LV = Planilha["TABELA PADRAO LV - LESTE-OESTE"]
    Aba_SIAGO = SIAGO.active

    for celulaSIAGO in Aba_SIAGO['L']:  
        linhaSIAGO = celulaSIAGO.row
        verificação = str(Aba_SIAGO["L{}".format(linhaSIAGO)].value)

        if "LINHA VIVA" in verificação:
            codigoSIAGO = str(Aba_SIAGO["F{}".format(linhaSIAGO)].value)
            serviço = str(Aba_SIAGO["G{}".format(linhaSIAGO)].value)
            qtdserviço = str(Aba_SIAGO["H{}".format(linhaSIAGO)].value)
            
            for celulaLV in LV['B']:
                linhaLV = celulaLV.row
                codigoLV = str(LV["B{}".format(linhaLV)].value)

                if codigoLV == codigoSIAGO:
                    if serviço == "R":
                        LV["G{}".format(linhaLV)] = qtdserviço
                    
                    elif codigoLV == "1052":
                        LV["E{}".format(linhaLV)] = qtdserviço
                    
                    elif codigoLV == "1053":
                        LV["E{}".format(linhaLV)] = qtdserviço
                    
                    elif codigoLV == "1056":
                        LV["E{}".format(linhaLV)] = qtdserviço

                    elif codigoLV == "1057":
                        LV["E{}".format(linhaLV)] = qtdserviço

                    elif codigoLV == "1058":
                        LV["E{}".format(linhaLV)] = qtdserviço

                    elif codigoLV == "1059":
                        LV["E{}".format(linhaLV)] = qtdserviço

                    else:
                        LV["D{}".format(linhaLV)] = qtdserviço

                else:
                    print("Aguarde... Escrevendo...")
        
        else:
            codigoSIAGO = str(Aba_SIAGO["F{}".format(linhaSIAGO)].value)

            if codigoSIAGO == "1119":
                codigoSIAGO = "4633"
                print(codigoSIAGO)
            elif codigoSIAGO == "1120":
                codigoSIAGO = "4632"
                print(codigoSIAGO)
            else:
                codigoSIAGO = codigoSIAGO
                
            serviço = str(Aba_SIAGO["G{}".format(linhaSIAGO)].value)
            qtdserviço = str(Aba_SIAGO["H{}".format(linhaSIAGO)].value)
            
            for celulaLM in LM['B']:
                linhaLM = celulaLM.row
                codigoLM = str(LM["B{}".format(linhaLM)].value)

                if codigoLM == codigoSIAGO:
                    if serviço == "R":
                        LM["G{}".format(linhaLM)] = qtdserviço

                    elif codigoLM == "1050":
                        LM["E{}".format(linhaLM)] = qtdserviço
                    
                    elif codigoLM == "1051":
                        LM["E{}".format(linhaLM)] = qtdserviço
                    
                    elif codigoLM == "1054":
                        LM["E{}".format(linhaLM)] = qtdserviço

                    elif codigoLM == "1055":
                        LM["E{}".format(linhaLM)] = qtdserviço

                    elif codigoLM == "1058":
                        LM["E{}".format(linhaLM)] = qtdserviço

                    elif codigoLM == "1059":
                        LM["E{}".format(linhaLM)] = qtdserviço
                    
                    else:
                        LM["D{}".format(linhaLM)] = qtdserviço

                else:
                    print("Aguarde... Escrevendo...")

    print("Planilha finalizada e salva.")

    Planilha.save("C:\\Users\\{}\\Documents\\Medição\\{}\\Mão de Obra {}.xlsx".format(usuario,n_Obra,n_Obra))

main()
